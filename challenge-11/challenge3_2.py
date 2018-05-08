from openpyxl import load_workbook, Workbook
import datetime

# 读取所有文件
wb = load_workbook('courses.xlsx')
students_sheet = wb['students']
time_sheet = wb['time']

def combine():
    # 创建combine表
    combine_sheet = wb.create_sheet(title='combine')

    # 写表头
    combine_sheet.append(['创建时间', '课程名称', '学习人数', '学习时间'])

    # 从students sheet中读取数据并写入combinesheet中
    for stu in students_sheet.values:
        if stu[2] != '学习人数':
            # 遍历time sheet 匹配学习时间数据
            for time in time_sheet.values:
                if time[1] == stu[1]:
                    # 将数据写入新sheet中
                    combine_sheet.append(list(stu) + [time[2]])
    # 覆盖保存原文件
    wb.save('courses.xlsx')

def split():
    # 载入combine sheet
    combine_sheet = wb['combine']
    # 创建list存储课程年份数据
    split_year = []

    # 遍历combine sheet获取年份数据
    for item in combine_sheet.values:
        if item[0] != '创建时间':
            split_year.append(item[0].strftime("%Y"))
    split_year = list(set(split_year))
    
    # 创建年份对应的表
    for year in split_year:
        wb_temp = Workbook()
        # 移除创建表格时默认生成的sheet
        wb_temp.remove(wb_temp.active)
        # 创建名称为年份的sheet
        ws = wb_temp.create_sheet(title= year)
        for data in combine_sheet.values:
            # 注意在进行时间字符串化时一定要注意表头，否则将会报错，因为表头不是datetime格式的，不具备strftime方法
            if data[0] != '创建时间':
                if data[0].strftime("%Y") == year:
                    ws.append(list(data))
        wb_temp.save('{}.xlsx'.format(year))

# 执行
if __name__ == '__main__':
    combine()
    split()